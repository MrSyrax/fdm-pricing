import os
from openpyxl import load_workbook

f = os.path.join(os.path.dirname(__file__), 'FDM_Print_Prices.xlsx')

if not os.path.exists(f):
    print('No file found yet. Save a quote first.')
else:
    wb = load_workbook(f)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) <= 1:
        print('No quotes saved yet.')
    else:
        fmt = '{:<12} {:<30} {:>10} {:>12} {:>11}'
        print(fmt.format('Date', 'Product', 'Plain', '+Primer', '+Paint'))
        print('-' * 78)
        for row in rows[1:]:
            print(fmt.format(*[str(v) if v is not None else '' for v in row[:5]]))
        print()
        print(f'Total quotes: {len(rows) - 1}')
